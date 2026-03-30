from datetime import datetime, timedelta
from openpyxl import Workbook


def now_iso():
    return datetime.now().replace(microsecond=0).isoformat()


def slugify(text: str) -> str:
    return (
        text.lower()
        .replace("&", "dan")
        .replace("/", "-")
        .replace(" ", "-")
        .replace("--", "-")
    )


def append_rows(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append(row)


def main():
    ts = now_iso()
    wb = Workbook()
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    categories = [
        ("CAT-001", "Bahasa", "Program bahasa tatap muka", "languages", 1),
        ("CAT-002", "Komputer", "Pelatihan komputer praktis", "monitor", 2),
        ("CAT-003", "Desain", "Kelas desain kreatif", "palette", 3),
        ("CAT-004", "Digital Marketing", "Strategi pemasaran digital", "megaphone", 4),
        ("CAT-005", "Public Speaking", "Komunikasi dan presentasi", "mic", 5),
        ("CAT-006", "Programming", "Kelas coding dan software", "code", 6),
        ("CAT-007", "Office Skills", "Skill administrasi perkantoran", "briefcase", 7),
        ("CAT-008", "Keuangan", "Akuntansi dan finansial", "wallet", 8),
        ("CAT-009", "Kecantikan", "Skill beauty dan makeup", "sparkles", 9),
        ("CAT-010", "Kuliner", "Program praktik kuliner", "chef-hat", 10),
        ("CAT-011", "Teknik", "Pelatihan skill teknis", "wrench", 11),
        ("CAT-012", "Anak & Remaja", "Program pelajar dan remaja", "users", 12),
    ]

    ws = wb.create_sheet("course_categories")
    append_rows(
        ws,
        [
            "id",
            "name",
            "slug",
            "description",
            "icon",
            "sort_order",
            "is_active",
            "created_at",
            "updated_at",
        ],
        [
            [c[0], c[1], slugify(c[1]), c[2], c[3], c[4], True, ts, ts]
            for c in categories
        ],
    )

    courses_seed = [
        (
            "Kursus Bahasa Inggris Intensif",
            "CAT-001",
            "Pemula",
            "8 minggu",
            16,
            1800000,
            1500000,
        ),
        (
            "Kelas Desain Grafis Tatap Muka",
            "CAT-003",
            "Pemula",
            "6 minggu",
            12,
            2200000,
            1900000,
        ),
        (
            "Pelatihan Microsoft Excel untuk Kerja",
            "CAT-007",
            "Pemula",
            "4 minggu",
            8,
            1450000,
            1200000,
        ),
        (
            "Bootcamp Digital Marketing Offline",
            "CAT-004",
            "Menengah",
            "10 minggu",
            20,
            3500000,
            2990000,
        ),
        (
            "Kelas Public Speaking Profesional",
            "CAT-005",
            "Pemula",
            "5 minggu",
            10,
            1600000,
            1350000,
        ),
        (
            "Kursus Programming Dasar",
            "CAT-006",
            "Pemula",
            "8 minggu",
            16,
            2400000,
            2100000,
        ),
        (
            "Workshop Akuntansi UMKM",
            "CAT-008",
            "Pemula",
            "3 minggu",
            6,
            1300000,
            1100000,
        ),
        ("Kelas Barista Pemula", "CAT-010", "Pemula", "4 minggu", 8, 2800000, 2500000),
        ("Kelas Make Up Basic", "CAT-009", "Pemula", "4 minggu", 8, 2100000, 1800000),
        (
            "Kursus Komputer Administrasi Perkantoran",
            "CAT-007",
            "Pemula",
            "6 minggu",
            12,
            1700000,
            1450000,
        ),
        (
            "Pelatihan AutoCAD Teknik Sipil",
            "CAT-011",
            "Menengah",
            "7 minggu",
            14,
            3200000,
            2890000,
        ),
        (
            "Kelas Coding untuk Remaja",
            "CAT-012",
            "Pemula",
            "6 minggu",
            12,
            1900000,
            1650000,
        ),
    ]

    courses = []
    for i, c in enumerate(courses_seed, start=1):
        title, cat_id, level, duration, sessions, price, promo = c
        cid = f"CRS-{i:03d}"
        courses.append(
            [
                cid,
                cat_id,
                title,
                slugify(title),
                f"{title} dengan praktik langsung dan pendampingan mentor.",
                f"{title} dirancang untuk hasil belajar aplikatif dengan kombinasi teori inti dan simulasi kasus nyata.",
                "Mampu menerapkan skill pada studi kasus nyata.",
                "Kurikulum terstruktur, mentor berpengalaman, dan evaluasi progres.",
                "Siswa, mahasiswa, fresh graduate, karyawan, profesional.",
                level,
                duration,
                sessions,
                "offline",
                "Komitmen hadir sesuai jadwal dan membawa perangkat jika dibutuhkan.",
                "Modul belajar, praktik kelas, konsultasi mentor.",
                True,
                price,
                promo,
                f"https://images.unsplash.com/photo-1523240795612-9a054b0db644?w=1200&sig={i}",
                i <= 4,
                True,
                title,
                f"Program {title} tatap muka untuk peningkatan skill praktis.",
                ts,
                ts,
            ]
        )

    ws = wb.create_sheet("courses")
    append_rows(
        ws,
        [
            "id",
            "category_id",
            "title",
            "slug",
            "short_description",
            "full_description",
            "learning_outcomes",
            "benefits",
            "target_participants",
            "level",
            "duration_text",
            "total_sessions",
            "format_type",
            "requirements",
            "facilities_included",
            "certificate_available",
            "price",
            "promo_price",
            "thumbnail",
            "featured",
            "published",
            "seo_title",
            "seo_description",
            "created_at",
            "updated_at",
        ],
        courses,
    )

    instructors_seed = [
        ("Dwi Pratama", "English Communication", 8, "CELTA, TOEFL ITP"),
        ("Rina Aulia", "Graphic Design", 7, "Adobe Certified Professional"),
        ("Arif Nugroho", "Data & Excel", 10, "MOS Expert"),
        ("Fajar Ramadhan", "Digital Marketing", 9, "Meta Blueprint"),
        ("Nadia Putri", "Public Speaking", 6, "BNSP Trainer"),
        ("Bagas Saputra", "Programming", 8, "AWS Certified Cloud Practitioner"),
        ("Mila Anggraini", "Barista & F&B", 5, "SCA Foundation"),
        ("Salsa Wibowo", "Makeup Artist", 6, "MUA Professional Certification"),
    ]
    instructors = []
    for i, ins in enumerate(instructors_seed, start=1):
        name, spec, years, cert = ins
        instructors.append(
            [
                f"INS-{i:03d}",
                name,
                slugify(name),
                f"https://images.unsplash.com/photo-1544005313-94ddf0286df2?w=800&sig={i}",
                spec,
                f"{name} adalah instruktur {spec} dengan pendekatan praktis dan mentoring intensif.",
                years,
                cert,
                i <= 4,
                True,
                ts,
                ts,
            ]
        )

    ws = wb.create_sheet("instructors")
    append_rows(
        ws,
        [
            "id",
            "full_name",
            "slug",
            "photo",
            "specialization",
            "bio",
            "experience_years",
            "certifications",
            "featured",
            "is_active",
            "created_at",
            "updated_at",
        ],
        instructors,
    )

    branches_seed = [
        ("Jakarta Selatan", "Jakarta", "Jl. TB Simatupang No. 88"),
        ("Bandung", "Bandung", "Jl. Buah Batu No. 120"),
        ("Surabaya", "Surabaya", "Jl. Darmo No. 45"),
        ("Yogyakarta", "Yogyakarta", "Jl. Kaliurang KM 5"),
        ("Bekasi", "Bekasi", "Jl. Ahmad Yani No. 21"),
    ]
    branches = []
    for i, b in enumerate(branches_seed, start=1):
        name, city, address = b
        branches.append(
            [
                f"BR-{i:03d}",
                name,
                city,
                address,
                "https://maps.google.com",
                f"0812000000{i:02d}",
                f"{slugify(name)}@kursusoffline.id",
                "Senin - Sabtu 09:00-20:00",
                True,
                ts,
                ts,
            ]
        )

    ws = wb.create_sheet("branches")
    append_rows(
        ws,
        [
            "id",
            "name",
            "city",
            "address",
            "maps_url",
            "phone",
            "email",
            "operating_hours",
            "is_active",
            "created_at",
            "updated_at",
        ],
        branches,
    )

    schedules = []
    base_date = datetime(2026, 4, 6)
    for i in range(1, 21):
        course_id = courses[(i - 1) % len(courses)][0]
        branch_id = branches[(i - 1) % len(branches)][0]
        instructor_id = instructors[(i - 1) % len(instructors)][0]
        start = base_date + timedelta(days=i * 2)
        end = start + timedelta(days=30)
        quota = 20 + (i % 3) * 5
        booked = 6 + (i % 9)
        available = max(quota - booked, 0)
        status = "open" if available > 5 else ("limited" if available > 0 else "full")
        schedules.append(
            [
                f"SCH-{i:03d}",
                course_id,
                branch_id,
                instructor_id,
                f"BATCH-{i:02d}",
                start.date().isoformat(),
                end.date().isoformat(),
                "Weekday" if i % 2 == 0 else "Weekend",
                "19:00-21:00" if i % 2 == 0 else "09:00-12:00",
                quota,
                booked,
                available,
                (start - timedelta(days=5)).date().isoformat(),
                status,
                "Jadwal reguler dengan praktik intensif.",
                ts,
                ts,
            ]
        )

    ws = wb.create_sheet("schedules")
    append_rows(
        ws,
        [
            "id",
            "course_id",
            "branch_id",
            "instructor_id",
            "batch_code",
            "start_date",
            "end_date",
            "class_days",
            "class_time",
            "quota",
            "booked_seats",
            "available_seats",
            "registration_deadline",
            "status",
            "notes",
            "created_at",
            "updated_at",
        ],
        schedules,
    )

    leads_names = [
        "Andi Saputra",
        "Siti Nurhaliza",
        "Budi Santoso",
        "Rina Marlina",
        "Fajar Hidayat",
        "Nisa Amalia",
        "Kevin Pratama",
        "Aulia Putri",
        "Rizki Maulana",
        "Dina Kurnia",
        "Farhan Akbar",
        "Maya Lestari",
        "Yusuf Ramadhan",
        "Ayu Laksmi",
        "Ilham Nugraha",
    ]
    statuses = [
        "new",
        "contacted",
        "follow_up",
        "confirmed",
        "paid",
        "enrolled",
        "cancelled",
    ]
    sources = [
        "instagram",
        "google",
        "friend",
        "tiktok",
        "corporate",
        "facebook",
        "website",
    ]
    backgrounds = [
        "Mahasiswa",
        "Karyawan",
        "Fresh Graduate",
        "UMKM Owner",
        "Profesional",
    ]
    registrations = []
    for i in range(1, 16):
        registrations.append(
            [
                f"REG-{i:03d}",
                courses[(i - 1) % len(courses)][0],
                schedules[(i - 1) % len(schedules)][0],
                leads_names[i - 1],
                f"08123{100000 + i}",
                f"lead{i}@mail.com",
                18 + (i % 12),
                backgrounds[i % len(backgrounds)],
                branches[(i - 1) % len(branches)][0],
                "Ingin jadwal yang tidak bentrok dengan kerja/kuliah.",
                sources[i % len(sources)],
                statuses[i % len(statuses)],
                f"staff-{(i % 3) + 1}",
                "Follow-up awal oleh tim admin.",
                ts,
                ts,
            ]
        )

    ws = wb.create_sheet("registrations")
    append_rows(
        ws,
        [
            "id",
            "course_id",
            "schedule_id",
            "full_name",
            "phone",
            "email",
            "age",
            "occupation_or_background",
            "preferred_branch_id",
            "notes",
            "lead_source",
            "status",
            "assigned_to",
            "follow_up_notes",
            "created_at",
            "updated_at",
        ],
        registrations,
    )

    testimonial_texts = [
        "Mentornya komunikatif dan materi langsung bisa dipakai di kerja.",
        "Kelasnya terstruktur, ada praktik dan feedback tiap sesi.",
        "Sangat membantu untuk upgrade skill sebelum apply kerja.",
        "Fasilitas nyaman dan jadwalnya fleksibel.",
        "Setelah ikut kelas, saya lebih percaya diri presentasi.",
        "Materi digital marketing sangat aplikatif untuk UMKM.",
        "Instruktur sabar dan detail saat membimbing praktik.",
        "Admin responsif, proses registrasi cepat dan jelas.",
        "Syllabus jelas dari dasar sampai project akhir.",
        "Worth it untuk investasi skill jangka panjang.",
    ]
    testimonials = []
    for i in range(1, 11):
        testimonials.append(
            [
                f"TST-{i:03d}",
                leads_names[i - 1],
                branches[(i - 1) % len(branches)][2],
                courses[(i - 1) % len(courses)][0],
                4 + (i % 2),
                testimonial_texts[i - 1],
                f"https://images.unsplash.com/photo-1500648767791-00dcc994a43e?w=800&sig={i}",
                "",
                i <= 5,
                True,
                ts,
                ts,
            ]
        )

    ws = wb.create_sheet("testimonials")
    append_rows(
        ws,
        [
            "id",
            "student_name",
            "city",
            "course_id",
            "rating",
            "review",
            "photo",
            "video_url",
            "is_featured",
            "is_published",
            "created_at",
            "updated_at",
        ],
        testimonials,
    )

    gallery_categories = [
        "classroom",
        "activity",
        "facility",
        "certificate",
        "student-life",
    ]
    galleries = []
    for i in range(1, 13):
        galleries.append(
            [
                f"GAL-{i:03d}",
                f"Dokumentasi Kelas {i}",
                "image",
                f"https://images.unsplash.com/photo-1523240795612-9a054b0db644?w=1200&sig={100 + i}",
                gallery_categories[i % len(gallery_categories)],
                courses[(i - 1) % len(courses)][0],
                branches[(i - 1) % len(branches)][0],
                i <= 4,
                i,
                True,
                ts,
                ts,
            ]
        )

    ws = wb.create_sheet("galleries")
    append_rows(
        ws,
        [
            "id",
            "title",
            "media_type",
            "media_url",
            "category",
            "related_course_id",
            "related_branch_id",
            "is_featured",
            "sort_order",
            "is_published",
            "created_at",
            "updated_at",
        ],
        galleries,
    )

    faq_items = [
        (
            "Apakah kelas cocok untuk pemula?",
            "Ya, tersedia level pemula hingga lanjutan.",
        ),
        (
            "Apakah bisa pilih kelas weekend?",
            "Bisa, kami menyediakan batch weekday dan weekend.",
        ),
        (
            "Apakah ada sertifikat?",
            "Ya, untuk program tertentu tersedia sertifikat penyelesaian.",
        ),
        (
            "Bagaimana cara daftar?",
            "Pilih program, pilih jadwal, isi form registrasi, lalu tunggu konfirmasi admin.",
        ),
        (
            "Apakah tersedia cicilan?",
            "Ketersediaan skema pembayaran bergantung program dan periode promo.",
        ),
        (
            "Bisa konsultasi sebelum daftar?",
            "Bisa, hubungi admin melalui WhatsApp atau form kontak.",
        ),
        (
            "Apakah corporate training tersedia?",
            "Ya, kami melayani pelatihan in-house untuk perusahaan/UMKM.",
        ),
        (
            "Apakah bisa pindah jadwal?",
            "Bisa sesuai ketentuan kuota dan kebijakan program.",
        ),
        ("Apakah materi diberikan?", "Ya, peserta mendapatkan modul sesuai program."),
        (
            "Lokasi kelas di mana saja?",
            "Saat ini tersedia di beberapa kota utama sesuai data cabang.",
        ),
    ]
    faqs = []
    for i, faq in enumerate(faq_items, start=1):
        q, a = faq
        faqs.append([f"FAQ-{i:03d}", q, a, "general", i, True, ts, ts])

    ws = wb.create_sheet("faqs")
    append_rows(
        ws,
        [
            "id",
            "question",
            "answer",
            "category",
            "sort_order",
            "is_published",
            "created_at",
            "updated_at",
        ],
        faqs,
    )

    banners = []
    for i in range(1, 6):
        course = courses[(i - 1) % len(courses)]
        start = datetime(2026, 4, 1) + timedelta(days=i * 10)
        end = start + timedelta(days=20)
        banners.append(
            [
                f"BNR-{i:03d}",
                f"Promo {course[2]}",
                "Diskon terbatas untuk pendaftaran batch terdekat",
                f"https://images.unsplash.com/photo-1509062522246-3755977927d7?w=1400&sig={200 + i}",
                "Daftar Sekarang",
                f"/course-detail.html?slug={course[3]}",
                start.date().isoformat(),
                end.date().isoformat(),
                True,
                ts,
                ts,
            ]
        )

    ws = wb.create_sheet("banners")
    append_rows(
        ws,
        [
            "id",
            "title",
            "subtitle",
            "image",
            "cta_label",
            "cta_url",
            "start_date",
            "end_date",
            "is_active",
            "created_at",
            "updated_at",
        ],
        banners,
    )

    content_blocks = [
        [
            "CB-001",
            "home",
            "hero",
            "Belajar Skill Praktis Lewat Kelas Tatap Muka",
            "Program siap kerja untuk individu & corporate",
            "Konten hero utama",
            "",
            "Daftar Sekarang",
            "/checkout.html",
            1,
            True,
            "Web Kursus Offline",
            "Platform pelatihan tatap muka profesional",
            ts,
            ts,
        ],
        [
            "CB-002",
            "home",
            "featured_programs",
            "Program Unggulan",
            "Pilih kelas paling populer",
            "Section program unggulan",
            "",
            "Lihat Program",
            "/courses.html",
            2,
            True,
            "Program Unggulan Kursus",
            "Program pelatihan paling diminati",
            ts,
            ts,
        ],
        [
            "CB-003",
            "home",
            "why_us",
            "Kenapa Memilih Kami",
            "Praktis, terstruktur, terukur",
            "Keunggulan lembaga",
            "",
            "Konsultasi Program",
            "/contact.html",
            3,
            True,
            "Alasan Memilih Kursus Kami",
            "Keunggulan pelatihan tatap muka",
            ts,
            ts,
        ],
        [
            "CB-004",
            "home",
            "upcoming_schedules",
            "Jadwal Terdekat",
            "Batch baru setiap bulan",
            "Section jadwal",
            "",
            "Lihat Jadwal",
            "/schedules.html",
            4,
            True,
            "Jadwal Kelas Terdekat",
            "Info batch kelas offline",
            ts,
            ts,
        ],
        [
            "CB-005",
            "home",
            "instructors",
            "Instruktur Profesional",
            "Berpengalaman industri",
            "Highlight instruktur",
            "",
            "Lihat Instruktur",
            "/instructors.html",
            5,
            True,
            "Instruktur Kursus",
            "Profil mentor berpengalaman",
            ts,
            ts,
        ],
        [
            "CB-006",
            "home",
            "testimonials",
            "Testimoni Peserta",
            "Cerita hasil belajar nyata",
            "Highlight testimoni",
            "",
            "Lihat Testimoni",
            "/testimonials.html",
            6,
            True,
            "Testimoni Peserta",
            "Ulasan alumni kursus",
            ts,
            ts,
        ],
        [
            "CB-007",
            "about",
            "institution_profile",
            "Profil Lembaga",
            "Komitmen mutu pelatihan",
            "Konten profil",
            "",
            "Hubungi Admin",
            "/contact.html",
            1,
            True,
            "Tentang Lembaga",
            "Profil institusi pelatihan",
            ts,
            ts,
        ],
        [
            "CB-008",
            "about",
            "vision_mission",
            "Visi & Misi",
            "Membangun talenta siap kerja",
            "Konten visi misi",
            "",
            "Lihat Program",
            "/courses.html",
            2,
            True,
            "Visi Misi Kursus",
            "Arah pengembangan lembaga",
            ts,
            ts,
        ],
        [
            "CB-009",
            "contact",
            "contact_info",
            "Kontak & Cabang",
            "Tim siap membantu konsultasi",
            "Informasi kontak",
            "",
            "Chat WhatsApp",
            "https://wa.me/6281200000000",
            1,
            True,
            "Kontak Kursus Offline",
            "Info cabang dan konsultasi",
            ts,
            ts,
        ],
        [
            "CB-010",
            "home",
            "faq",
            "Pertanyaan Umum",
            "Jawaban seputar program",
            "FAQ ringkas",
            "",
            "Baca FAQ",
            "/faq",
            7,
            True,
            "FAQ Kursus",
            "Pertanyaan umum calon peserta",
            ts,
            ts,
        ],
    ]

    ws = wb.create_sheet("content_blocks")
    append_rows(
        ws,
        [
            "id",
            "page_slug",
            "block_key",
            "title",
            "subtitle",
            "content",
            "image_url",
            "cta_label",
            "cta_url",
            "sort_order",
            "is_published",
            "seo_title",
            "seo_description",
            "created_at",
            "updated_at",
        ],
        content_blocks,
    )

    admin_users = [
        [
            "ADM-001",
            "Super Admin",
            "superadmin@kursusoffline.id",
            "<HASHED_PASSWORD>",
            "super_admin",
            "",
            True,
            ts,
            ts,
            ts,
        ],
        [
            "ADM-002",
            "Admin Staff",
            "staff@kursusoffline.id",
            "<HASHED_PASSWORD>",
            "admin_staff",
            "",
            True,
            ts,
            ts,
            ts,
        ],
        [
            "ADM-003",
            "Branch Admin Jakarta",
            "jakarta.admin@kursusoffline.id",
            "<HASHED_PASSWORD>",
            "branch_admin",
            "BR-001",
            True,
            ts,
            ts,
            ts,
        ],
        [
            "ADM-004",
            "Marketing Team",
            "marketing@kursusoffline.id",
            "<HASHED_PASSWORD>",
            "marketing",
            "",
            True,
            ts,
            ts,
            ts,
        ],
        [
            "ADM-005",
            "Content Editor",
            "content@kursusoffline.id",
            "<HASHED_PASSWORD>",
            "content_editor",
            "",
            True,
            ts,
            ts,
            ts,
        ],
    ]
    ws = wb.create_sheet("admin_users")
    append_rows(
        ws,
        [
            "id",
            "name",
            "email",
            "password_hash",
            "role",
            "branch_id",
            "is_active",
            "last_login_at",
            "created_at",
            "updated_at",
        ],
        admin_users,
    )

    output_path = "dummy_web_kursus_offline_import.xlsx"
    wb.save(output_path)
    print(output_path)
    print("sheets:", len(wb.sheetnames))
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"- {name}: {ws.max_row - 1} rows")


if __name__ == "__main__":
    main()
